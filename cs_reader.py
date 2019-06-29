#!/usr/bin/python
# coding=utf-8

import types
import sys
import re 
import os
from openpyxl import load_workbook

filename = sys.argv[1].strip()
path_split = filename.split("/")
name_split = path_split[len(path_split)-1].split(".")
excel_name = name_split[0]
excel_name = excel_name.capitalize()

resultPath = "./Result/cs/"
if not os.path.isdir(resultPath):
    os.makedirs(resultPath)

configPath = "./Config"
if not os.path.isdir(configPath):
    os.makedirs(configPath)

wb = load_workbook(filename)
sheet_names = wb.get_sheet_names()
listSheet = wb.get_sheet_by_name(name=sheet_names[0]) #读第一个sheet
col = listSheet.get_highest_column()
row = listSheet.get_highest_row()
for i in range(row):
    cell = listSheet.cell(column=0,row=i)
    v_type = type(cell.value)
    if v_type is types.UnicodeType :
        value = cell.value.encode('utf-8')
    elif v_type is types.NoneType :
        value = ''
    else :
        value = str(cell.value)
    sheet_name = value.capitalize()
    dataSheet = wb.get_sheet_by_name(value)
    data_col = dataSheet.get_highest_column()
    data_row = dataSheet.get_highest_row()
    
    #====================================================================
    #输出资源文件，同时读出变量列表
    struct_name = []    #变量名
    struct_type = []    #变量类型 
    struct_desc = []    #变量注释
    rcs_path = configPath+"/"+excel_name+"_"+sheet_name+".txt"  #资源路径
    rcs_file = open(rcs_path,"w")
    for r in range(data_row):
        out = ""
        for c in range(data_col):
            data_cell = dataSheet.cell(column=c,row=r)
            data_type = type(data_cell.value)
            if data_type is types.UnicodeType :
                data_value = data_cell.value.encode('utf-8')
            elif data_type is types.NoneType :
                data_value = ''
            else :
                data_value = str(data_cell.value)
            if c != 0 or (c == 0 and data_cell.data_type == "s" ) :
                out += "#"
            out += data_value
            if r == 0:
                struct_name.append(data_value)
            if r == 1:
                struct_desc.append(data_value)
            if r == 2:
                struct_type.append(data_value)
        rcs_file.write(out+os.linesep)
    rcs_file.close()
    #====================================================================
    #输出成员变量，同时记录是否数组或者结构体
    config_name = excel_name+sheet_name+"Config"
    out_path = resultPath+excel_name+sheet_name+"ConfigTable.cs"
    out_file = open(out_path,"w")
    out_file.write("using System.Collections;"+os.linesep +"using System.Collections.Generic;"+os.linesep)
    out_file.write("public class "+config_name+os.linesep+"{"+os.linesep)

    i = 0;  #变量index
    type_num = {}       #变量是否数组标志，>1为数组   
    subStruct_name = {}  #key:结构体变量名，value：结构体类名
    subStruct_mem = {}  #key:结构体变量名，value：成员变量名数组
    subStruct_desc = {} #key:结构体变量名，value：成员变量注释数组
    subStruct_type = {} #key:结构体变量名，value：成员变量类型数组
    while i < len(struct_name):
        out_file.write("\tpublic\t")
        bSame = False
        bStruct = False
        num = 1;
        if struct_name[i].find("-") > 0:
            bStruct = True
            sub_name = struct_name[i].split("-")[0]
            sub_type = excel_name + sheet_name + sub_name.capitalize() + "Config"
            subStruct_name[sub_name] = sub_type
            out_file.write(sub_type+"[]\t" + sub_name + ";"+os.linesep)
            mems = []
            descs = []
            typesL = []
            while struct_name[i].find("-") > 0:
                sub_mem = struct_name[i].split("-")[1]
                if (not sub_mem in mems):
                    mems.append(sub_mem)
                    descs.append(struct_desc[i])
                    typesL.append(struct_type[i])
                    type_num[sub_name] = 1
                else:
                    if (sub_mem == mems[-1]):
                        type_num[sub_name] = type_num[sub_name] + 1
                type_num[struct_name[i]] = 1
                i = i + 1
            i = i - 1
            subStruct_mem[sub_name] = mems
            subStruct_desc[sub_name] = descs
            subStruct_type[sub_name] = typesL
        else :
            out_file.write(struct_type[i])
            while (i + 1 < len(struct_name)) and (struct_name[i] == struct_name[i+1]):
                if bSame == False:
                    bSame = True;
                    out_file.write("[]")
                num = num+1
                i = i+1
        if not bStruct:
            out_file.write("\t"+struct_name[i]+";\t//"+struct_desc[i]+os.linesep)
        type_num[struct_name[i]] = num;
        i = i+1

    #====================================================================
    #声明所有结构体
    out_file.write("}"+os.linesep)
    for struct in subStruct_name:
        out_file.write(os.linesep+"public class " + subStruct_name[struct] + os.linesep)
        out_file.write("{"+os.linesep)
        mems = subStruct_mem[struct]
        descs = subStruct_desc[struct]
        typesL = subStruct_type[struct] 
        for i in range(len(mems)):
            out_file.write("\tpublic " + typesL[i] + "\t" +mems[i] + ";\t//" + descs[i] + os.linesep)
        out_file.write("}"+os.linesep)
    out_file.write(os.linesep)

    #====================================================================
    #数据加载方法
    out_file.write("public class "+config_name+"Table"+os.linesep+"{"+os.linesep)
    out_file.write("\tDictionary<int, "+config_name+"> m_configs = new Dictionary<int,"+config_name+">();"+os.linesep)
    out_file.write("\tpublic Dictionary<int, "+config_name+"> configs"+os.linesep+"\t{"+os.linesep+"\t\tget {"+os.linesep+"\t\t\treturn m_configs;"+os.linesep+"\t\t}"+os.linesep+"\t}"+os.linesep+os.linesep)
    out_file.write("\tpublic "+config_name+" GetConfigById(int cid)"+os.linesep+"\t{"+os.linesep+"\t\tif(m_configs.ContainsKey(cid)){"+os.linesep+"\t\t\treturn m_configs[cid];"+os.linesep+"\t\t}"+os.linesep+"\t\treturn null;"+os.linesep+"\t}"+os.linesep+""+os.linesep+"")
    out_file.write("\tpublic object ConfigProcess(string[] row)"+os.linesep+"\t{"+os.linesep)
    out_file.write("\t\tif (row.Length < ")
    out_file.write("%d)"%len(struct_type))
    out_file.write(os.linesep)
    out_file.write("\t\t{"+os.linesep+"\t\t\treturn null;"+os.linesep+"\t\t}"+os.linesep+""+os.linesep+"\t\tRowHelper rh = new RowHelper(row);"+os.linesep+"")
    out_file.write("\t\t"+config_name+" rec = new "+config_name+"();"+os.linesep+""+os.linesep+"")
    count = 0
    
    now_struct = ''  #当前处理的结构体变量
    now_index = 0   #当前处理的成员变量下标
    for i in range(len(struct_type)):
        if struct_name[i].find("-") > 0:
            sub_name = struct_name[i].split("-")[0]
            sub_mem = struct_name[i].split("-")[1]
            if not sub_name == now_struct :
                now_struct = sub_name
                now_index = 0
                mem_num = len(subStruct_mem[sub_name])
                out_file.write("\t\trec." + sub_name + " = new " + subStruct_name[sub_name] +"[%d];"%type_num[sub_name] +""+os.linesep+"")
                out_file.write("\t\tfor (int i = 0;i < %d;i ++) {"%type_num[sub_name])
                out_file.write(os.linesep)
                out_file.write("\t\t\trec." + sub_name + "[i] = new " + subStruct_name[sub_name] + "();"+os.linesep+"")
                out_file.write("\t\t}"+os.linesep+"")
            out_file.write("\t\trec." + sub_name + "[%d]."%now_index + sub_mem + " = ")
            if sub_mem == subStruct_mem[sub_name][-1]:
                now_index = now_index + 1
        else:
            if count < type_num[struct_name[i]] and type_num[struct_name[i]] > 1:
                if count == 0:
                    out_file.write("\t\trec."+struct_name[i]+" = new "+struct_type[i]+"[")
                    out_file.write("%d"%type_num[struct_name[i]])
                    out_file.write("];"+os.linesep+"")
                out_file.write("\t\trec."+struct_name[i]+"[")
                out_file.write("%d"%count)
                out_file.write("] = ")
                count = count + 1
            else:
                count = 0
                if type_num[struct_name[i]] == 1:
                    out_file.write("\t\trec."+struct_name[i]+" = ")
        if struct_type[i] == "int":
            out_file.write("CSVUtility.ToInt(rh.Read());")
        elif struct_type[i] == "float":
            out_file.write("CSVUtility.ToFloat(rh.Read());")
        elif struct_type[i] == "string":
            out_file.write("rh.Read();")
        out_file.write("\t//"+struct_desc[i]+""+os.linesep+"")
    out_file.write("\t\treturn rec;"+os.linesep+"\t}"+os.linesep+""+os.linesep+"\tpublic void Load()"+os.linesep+"\t{"+os.linesep+"\t\tCSVReader reader = new CSVReader();"+os.linesep+"")
    out_file.write("\t\treader.LoadText(\"Config/"+excel_name+"_"+sheet_name+".txt\", 3);"+os.linesep+"")
    out_file.write("\t\tint rows = reader.GetRowCount();"+os.linesep+"\t\tfor (int r = 0; r < rows; ++r)"+os.linesep+"\t\t{"+os.linesep+"\t\t\tstring[] row = reader.GetRow(r);"+os.linesep+"")
    out_file.write("\t\t\t"+config_name+" ac = ConfigProcess(row) as "+config_name+";"+os.linesep+"")
    out_file.write("\t\t\tconfigs.Add(ac."+struct_name[0]+", ac);"+os.linesep+"\t\t}"+os.linesep+"\t}"+os.linesep+"}"+os.linesep+"")
    out_file.close();





