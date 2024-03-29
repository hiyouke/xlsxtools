#!/usr/bin/python
# coding=utf-8

import types
import sys
import os
from openpyxl import load_workbook

# 类成员
class MemberDesc:
    def __init__(self, name="", type="", desc="", count=1, isobj = False):
        self.name = 'm_' + name # 变量名
        self.type = type # 变量类型
        self.desc = desc # 变量注释
        self.count = count # 为1时不是数组
        self.isobj = isobj # 是否oc对象

# 类结构体
class ClassDesc:
    def __init__(self, name=""):
        # 类名
        self.name = name
        self.members = []
        self.member_names = []
        # 构造这类要用到的字符串数
        self.init_count = 0


def get_type_str(s_type):
    str_ret = ''
    if s_type == "int":
        str_ret = '[rows[index++] intValue]'
    elif s_type == "float":
        str_ret = '[rows[index++] floatValue]'
    elif s_type == "NSString":
        str_ret = 'rows[index++]'
    else:
        print('unsupoprt type : ' + s_type)
    return str_ret


def makeClasses(excel_name, sheet_name, type_map, isobj_map, class_member_names, class_member_descs, class_member_types):
    i = 0
    class_name = classNameWithTableName(excel_name, sheet_name)
    main_class = ClassDesc(class_name)
    main_class.init_count = len(class_member_names)
    find_classes = [main_class]
    while i < len(class_member_names):
        member_name = class_member_names[i]
        member_desc = class_member_descs[i]
        member_type = class_member_types[i]
        if member_name.find("-") > 0:
            member_count = 1
            sub_class_mark = member_name.split("-")[0]
            sub_class_name = sheet_name + sub_class_mark.capitalize()
            sub_class_find = False
            sub_class_from = i
            sub_class_member_names = []
            i = i + 1
            while (i < len(class_member_names)) and (class_member_names[i].find("-") > 0):
                sub_class_mark2 = class_member_names[i].split("-")[0]
                sub_class_member_names.append(class_member_names[i][len(sub_class_mark2) + 1:])
                if sub_class_mark == sub_class_mark2:
                    if member_name == class_member_names[i]:
                        if sub_class_find == False:
                            sub_class_find = True
                            t_class = makeClasses(excel_name, sub_class_name, type_map, isobj_map,sub_class_member_names, class_member_descs[sub_class_from:i], class_member_types[sub_class_from:i])
                            find_classes = find_classes + t_class
                        member_count = member_count + 1
                else:
                    break
                i = i + 1
            if sub_class_find == False:
                t_class = makeClasses(excel_name, sub_class_name, type_map, isobj_map, sub_class_member_names, class_member_descs[sub_class_from:i], class_member_types[sub_class_from:i])
                find_classes = find_classes + t_class
            i = i - 1
            sub_class_name = classNameWithTableName(excel_name, sub_class_name)
            sub_member = MemberDesc(sub_class_mark + "s", sub_class_name, "", member_count, True)
            main_class.members.append(sub_member)
            main_class.member_names.append(member_name)
        else:
            # 先判断是否数组
            if member_name in main_class.member_names:
                member_index = main_class.member_names.index(member_name)
                member = main_class.members[member_index]
                member.count = member.count + 1
            else:
                sub_member = MemberDesc(member_name, type_map[member_type], member_desc, 1, isobj_map[member_type])
                main_class.members.append(sub_member)
                main_class.member_names.append(member_name)
        i = i + 1
    return find_classes


def writeHeader(result_path, class_list, excel_name, sheet_name):
    out_path = result_path + "/" + excel_name + sheet_name + "ConfigTable.h"
    out_file = open(out_path, "w")
    out_string = "#import <Foundation/Foundation.h>"+os.linesep+""
    out_string += "#include <vector>"+os.linesep+""+os.linesep+""
    out_string += "using namespace std;"+os.linesep+""

    # class预定义
    for i in range(len(class_list)):
        curr_class = class_list[i]
        out_string += "@class " + curr_class.name + ";"+os.linesep+""
    # class定义
    for i in range(len(class_list)):
        curr_class = class_list[i]
        out_string += ""+os.linesep+"@interface " + curr_class.name + " : NSObject"+os.linesep+""+os.linesep+""
        for j in range(len(curr_class.members)):
            class_member = curr_class.members[j]
            if class_member.count > 1:
                out_string += "@property (nonatomic) vector<%s%s> %s; // count: %d"%(class_member.type, 
                    (class_member.isobj) and '*' or '', class_member.name, class_member.count)
                out_string += os.linesep
            else:
                out_string += "@property (nonatomic) %s%s %s;"%(class_member.type, (class_member.isobj) and '*' or '', 
                    class_member.name)
                out_string += os.linesep
        out_string += os.linesep
        out_string += "+(%s*)ConfigProcess:(NSArray*)rows;"%curr_class.name
        out_string += os.linesep
        out_string += ""+os.linesep+"@end"+os.linesep+""
    # table定义
    out_string += ""+os.linesep+"@interface " + classNameWithTableName(excel_name, sheet_name) + "Table" + " : NSObject"+os.linesep+""+os.linesep+""
    out_string += "+ (NSDictionary*)configs;"+os.linesep+""
    out_string += ""+os.linesep+"@end"+os.linesep+""
    # print(out_string)
    out_file.write(out_string)


def writeMMSource(result_path, class_list, normal_type_map, excel_name, sheet_name):
    out_path = result_path + "/" + excel_name + sheet_name + "ConfigTable.mm"
    out_file = open(out_path, "w")
    out_string = "#import \"" + excel_name + sheet_name + "ConfigTable.h\""+os.linesep+""
    # class定义
    for i in range(len(class_list)):
        curr_class = class_list[i]
        out_string += ""+os.linesep+"@implementation " + curr_class.name + ""+os.linesep+""+os.linesep+""
        for j in range(len(curr_class.members)):
            class_member = curr_class.members[j]
            out_string += "@synthesize " + class_member.name + ";"+os.linesep+""
        # 构造方法
        out_string += os.linesep
        out_string += "+(%s*)ConfigProcess:(NSArray*)rows"%curr_class.name
        out_string += os.linesep
        out_string += "{"+os.linesep+""
        out_string += "\tif (rows.count != %d) {" % curr_class.init_count
        out_string += (os.linesep+"\t\treturn nil;"+os.linesep+"\t}"+os.linesep+""+os.linesep)
        out_string += "\tint index = 0;"+os.linesep+""
        out_string += "\t%s* newInstance = [[%s alloc] init];" % (curr_class.name, curr_class.name)
        out_string += os.linesep+os.linesep
        for j in range(len(curr_class.members)):
            class_member = curr_class.members[j]
            tab_mark = ""
            index_mark = ""

            # print("class_member name : %s class_member count  : %d" % (class_member.name, class_member.count))

            if class_member.count > 1:
                if class_member.type in normal_type_map:
                    out_string += '''
    vector<%s> %s_vector;
    for (int i = 0; i < %d; i++) 
    {
        %s_vector.push_back(%s);
    }
    newInstance.%s = %s_vector;
    \n''' % (class_member.type, class_member.name, class_member.count, class_member.name,
             get_type_str(class_member.type), class_member.name, class_member.name)
                else:
                    member_class = findClassWithName(class_list, class_member.type)
                    out_string += '''
    vector<%s *> %s_vector;
    for (int i = 0; i < %d; i++) 
    {
        %s_vector.push_back([%s ConfigProcess:[rows subarrayWithRange:NSMakeRange(index, %d)]]);
        index += %d;
    }
    newInstance.%s = %s_vector;
    \n''' % (member_class.name, class_member.name, class_member.count, class_member.name,
             member_class.name, member_class.init_count, member_class.init_count,
             class_member.name, class_member.name)
            else:
                out_string += "\tnewInstance.%s = %s;\n" % (class_member.name, get_type_str(class_member.type))
        out_string += ""+os.linesep+"\treturn newInstance;"+os.linesep+"}"+os.linesep+""
        out_string += ""+os.linesep+"@end"+os.linesep+""

    # table定义
    out_string += '''
@implementation %sTable

+ (NSDictionary*)configs
{
    NSMutableDictionary* configs = [NSMutableDictionary dictionary];
    NSString * path = [[NSBundle mainBundle] pathForResource:@"Config/%s_%s.txt" ofType:nil];
    NSString* fileString = [NSString stringWithContentsOfFile:path encoding:NSUTF8StringEncoding error:nil];
    NSArray* lines = [fileString componentsSeparatedByString:@"\\r\\n"];
    for (int i = 3; i < lines.count; ++i) 
    {
        NSArray* line = [lines[i] componentsSeparatedByString:@"#"];
        if([line count] > 1)
        {
            %s%sConfig* config = [%s%sConfig ConfigProcess:line];
            [configs setObject:config forKey:[NSNumber numberWithInt:config.m_id]];
        }
    }
    return configs;
}

@end
''' % (classNameWithTableName(excel_name, sheet_name), excel_name, sheet_name, excel_name, sheet_name, excel_name, sheet_name)
    out_file.write(out_string)


def findClassWithName(class_list, name):
    for curr_class in class_list:
        if curr_class.name == name:
            return curr_class


# 结构名=文件名+表名+"Config"
def classNameWithTableName(excel_name, sheet_name):
    return excel_name + sheet_name + "Config";


def main():
    file_path = sys.argv[1].strip()
    path_split = file_path.split("/")
    name_split = path_split[len(path_split) - 1].split(".")
    excel_name = name_split[0].capitalize()

    normal_type_map = {"int": "int", "float": "float", "string": "NSString"}
    isobj_map = {"int": False, "float": False, "string": True}

    result_path = "./Result/oc/"
    if not os.path.isdir(result_path):
        os.makedirs(result_path)

    config_path = "./Config"
    if not os.path.isdir(config_path):
        os.makedirs(config_path)

    work_book = load_workbook(file_path)
    sheet_names = work_book.get_sheet_names()
    first_sheet = work_book.get_sheet_by_name(sheet_names[0]) # 读第一个sheet
    col = first_sheet.get_highest_column()
    row = first_sheet.get_highest_row()
    for i in range(row):
        cell = first_sheet.cell(column = 0, row = i)
        value_type = type(cell.value)
        if value_type is types.UnicodeType:
            cell_value = cell.value.encode('utf-8')
        elif value_type is types.NoneType:
            cell_value = ''
        else:
            cell_value = str(cell.value)
        sheet_name = cell_value.capitalize()
        data_sheet = work_book.get_sheet_by_name(cell_value)
        data_col = data_sheet.get_highest_column()
        data_row = data_sheet.get_highest_row()

        # ====================================================================
        # 输出资源文件，同时读出变量列表
        class_member_names = []
        class_member_descs = []
        class_member_types = []
        for r in range(data_row):
            for c in range(data_col):
                data_cell = data_sheet.cell(column=c, row=r)
                data_cell_type = type(data_cell.value)
                if data_cell_type is types.UnicodeType:
                    data_cell_value = data_cell.value.encode('utf-8')
                elif data_cell_type is types.NoneType:
                    data_cell_value = ''
                else:
                    data_cell_value = str(data_cell.value)
                if r == 0:
                    class_member_names.append(data_cell_value)
                if r == 1:
                    class_member_descs.append(data_cell_value)
                if r == 2:
                    class_member_types.append(data_cell_value)

        # ====================================================================
        # 构造类结构体
        find_classes = makeClasses(excel_name, sheet_name, normal_type_map, isobj_map, class_member_names, class_member_descs, class_member_types)
        writeHeader(result_path, find_classes, excel_name, sheet_name)
        writeMMSource(result_path, find_classes, normal_type_map, excel_name, sheet_name)


main()
