#!/usr/bin/python
# coding=utf-8

import sys
import os
from openpyxl import load_workbook
from xlsx_to_txt import get_cell_value


def get_main_struct(main_excel_name, main_sheet_name, main_struct_name, main_struct_desc, main_struct_type):
    str_main_struct = ""
    name_index = 0  # 变量index
    type_num = {}  # 变量是否数组标志，>1为数组
    sub_struct_name = {}  # key:结构体变量名，value：结构体类名
    sub_struct_mem = {}  # key:结构体变量名，value：成员变量名数组
    sub_struct_desc = {}  # key:结构体变量名，value：成员变量注释数组
    sub_struct_type = {}  # key:结构体变量名，value：成员变量类型数组
    while name_index < len(main_struct_name):
        b_same = False
        b_struct = False
        num = 1
        if main_struct_name[name_index].find("-") > 0:
            b_struct = True
            sub_name = main_struct_name[name_index].split("-")[0]
            sub_type = main_excel_name + main_sheet_name + sub_name.capitalize() + "Config"
            sub_struct_name[sub_name] = sub_type
            str_main_struct += ('\tvector<%s> %s;\n' % (sub_type, sub_name))
            mems = []
            descs = []
            typesL = []
            while main_struct_name[name_index].find("-") > 0:
                sub_mem = main_struct_name[name_index].split("-")[1]
                if (not sub_mem in mems):
                    mems.append(sub_mem)
                    descs.append(main_struct_desc[name_index])
                    typesL.append(main_struct_type[name_index])
                    type_num[sub_name] = 1
                else:
                    if (sub_mem == mems[-1]):
                        type_num[sub_name] = type_num[sub_name] + 1
                type_num[main_struct_name[name_index]] = 1
                name_index = name_index + 1
            name_index = name_index - 1
            sub_struct_mem[sub_name] = mems
            sub_struct_desc[sub_name] = descs
            sub_struct_type[sub_name] = typesL
        else:
            while (name_index + 1 < len(main_struct_name)) and \
                    (main_struct_name[name_index] == main_struct_name[name_index + 1]):
                if not b_same:
                    b_same = True
                    str_main_struct += ("\tvector<%s>" % main_struct_type[name_index])
                num = num + 1
                name_index = name_index + 1
            if 1 == num:
                str_main_struct += ('\t' + main_struct_type[name_index])
        if not b_struct:
            str_main_struct += (" " + main_struct_name[name_index] + ";//" + main_struct_desc[name_index] + "\n")
        type_num[main_struct_name[name_index]] = num
        name_index = name_index + 1
    str_main_struct += ("};" + "\n")
    return str_main_struct, type_num, sub_struct_name, sub_struct_mem, sub_struct_desc, sub_struct_type


def get_sub_struct(sub_struct_name, sub_struct_mem, sub_struct_desc, sub_struct_type):
    str_sub_struct = ""
    for struct in sub_struct_name:
        str_sub_struct += ("\n"+"struct " + sub_struct_name[struct] + "\n")
        str_sub_struct += ("{"+"\n")
        mems = sub_struct_mem[struct]
        descs = sub_struct_desc[struct]
        typesL = sub_struct_type[struct]
        for i in range(len(mems)):
            str_sub_struct += ("\t" + typesL[i] + "\t" +mems[i] + ";\t//" + descs[i] + "\n")
        str_sub_struct += ("};"+"\n")
    str_sub_struct += ("\n")
    return str_sub_struct


def get_type_str(s_type):
    str_ret = ''
    if s_type == "int":
        str_ret = 'stoi(fields[idx++])'
    elif s_type == "float":
        str_ret = 'stof(fields[idx++])'
    elif s_type == "string":
        str_ret = 'fields[idx++]'
    else:
        print('unsupoprt type : ' + s_type)
    return str_ret


def get_config_process(config_struct_type, config_type_num, config_struct_name, config_struct_desc, sub_struct_name, sub_struct_mem):
    str_config_process = '\n'
    count = 0
    now_struct = ''  #当前处理的结构体变量
    now_index = 0   #当前处理的成员变量下标
    for i_item in range(len(config_struct_type)):
        if config_struct_name[i_item].find("-") > 0:
            sub_name = config_struct_name[i_item].split("-")[0]
            sub_mem = config_struct_name[i_item].split("-")[1]
            if not sub_name == now_struct:
                now_struct = sub_name
                now_index = 0
                mem_num = len(sub_struct_mem[sub_name])
                str_config_process += ('''
    for (int i = 0; i < %d; i++) 
    {
        rec.sell.push_back(%s());
    }\n''' % (config_type_num[sub_name], sub_struct_name[sub_name]))
            str_config_process += ("\trec." + sub_name + "[%d]." % now_index + sub_mem +
                                   " = " + get_type_str(config_struct_type[i_item])+';')
            if sub_mem == sub_struct_mem[sub_name][-1]:
                now_index = now_index + 1
        else:
            if config_type_num[config_struct_name[i_item]] > 1 and count < config_type_num[config_struct_name[i_item]]:
                str_config_process += ('\trec.'+config_struct_name[i_item]+'.push_back('+get_type_str(config_struct_type[i_item])+");")
                count = count + 1
            else:
                count = 0
                str_config_process += ("\trec." + config_struct_name[i_item] + " = "
                                       + get_type_str(config_struct_type[i_item])+";")
        str_config_process += ("//" + config_struct_desc[i_item] + '\n')
    return str_config_process


filename = sys.argv[1].strip()
path_split = filename.split("/")
name_split = path_split[len(path_split)-1].split(".")
excel_name = name_split[0]
excel_name = excel_name.capitalize()

resultPath = "./Result/cpp/"
if not os.path.isdir(resultPath):
    os.makedirs(resultPath)

wb = load_workbook(filename)
sheet_names = wb.get_sheet_names()
listSheet = wb.get_sheet_by_name(name=sheet_names[0]) #读第一个sheet
col = listSheet.get_highest_column()
row = listSheet.get_highest_row()
for i in range(row):
    cell = listSheet.cell(column=0,row=i)
    value = get_cell_value(cell)
    sheet_name = value.capitalize()
    dataSheet = wb.get_sheet_by_name(value)
    data_col = dataSheet.get_highest_column()
    data_row = dataSheet.get_highest_row()
    
    #====================================================================
    #输出资源文件，同时读出变量列表
    struct_name = []    #变量名
    struct_type = []    #变量类型 
    struct_desc = []    #变量注释
    for r in range(data_row):
        for c in range(data_col):
            data_cell = dataSheet.cell(column=c,row=r)
            data_value = get_cell_value(data_cell)
            if r == 0:
                struct_name.append(data_value)
            if r == 1:
                struct_desc.append(data_value)
            if r == 2:
                struct_type.append(data_value)

    #====================================================================
    #输出成员变量，同时记录是否数组或者结构体
    config_name = excel_name+sheet_name+"Config"
    class_name = excel_name+sheet_name+"ConfigTable"
    head_file_name = excel_name + sheet_name + "ConfigTable.h"
    head_out_path = resultPath + head_file_name
    head_out_file = open(head_out_path, "w")
    head_out_file.write("#pragma once\n#include<string>\n#include<vector>\n#include<map>\n")
    head_out_file.write("\nusing namespace std;\n")

    # ====================================================================
    # 生成主结构体
    main_struct, main_type_num, subStruct_name, subStruct_mem, subStruct_desc, subStruct_type = get_main_struct(excel_name, sheet_name,struct_name,struct_desc,struct_type)

    # ====================================================================
    # 声明所有子结构体
    head_out_file.write(get_sub_struct(subStruct_name, subStruct_mem, subStruct_desc, subStruct_type))

    # ====================================================================
    # 写入主结构体
    head_out_file.write("\nstruct " + config_name + "\n" + "{" + "\n")
    head_out_file.write(main_struct)

    # ====================================================================
    # 数据加载方法
    head_out_file.write("\nclass "+config_name+"Table"+"\n"+"{"+"\n")
    head_out_file.write("private:\n\tmap<int, "+config_name+"> m_configs;"+"\n")
    head_out_file.write("public:\n\tmap<int, "+config_name+"> & GetConfigs();"+"\n")
    head_out_file.write("\t"+config_name+" GetConfigById(int cid);"+"\n")
    head_out_file.write("\tvoid Load();\n" + "\n")
    head_out_file.write('''\tbool ConfigProcess(vector<string> fields, %s & rec);
};''' % (config_name,))

    # ====================================================================
    # CPP文件
    cpp_out_path = resultPath+excel_name+sheet_name+"ConfigTable.cpp"
    cpp_out_file = open(cpp_out_path, "w")
    cpp_out_file.write('#include "stdafx.h"\n#include "'+head_file_name+'"\n#include "CSVReader.h"\n')
    cpp_out_file.write('''
map<int, %s> & %s::GetConfigs()
{
    return m_configs;
}
''' % (config_name, class_name))

    cpp_out_file.write('''
void %s::Load()
{
    CSVReader reader;
    reader.LoadText("Config/%s.txt", 3);
    int rows = reader.GetRowCount();
    for (int r = 0; r < rows; ++r)
    {
        vector<string> row = reader.GetRow(r);
        %s ac;
        ConfigProcess(row, ac);
        m_configs.insert(std::make_pair(ac.id, ac));
    }
}\n''' % (class_name, excel_name+'_'+sheet_name, config_name))

    cpp_out_file.write('''
bool %s::ConfigProcess(vector<string> fields, %s & rec)
{
    if (fields.size() < %d)
    {
        return false;
    }
    int idx = 0;
''' % (class_name, config_name, len(struct_type)))

    str_process = get_config_process(struct_type, main_type_num, struct_name, struct_desc, subStruct_name, subStruct_mem)
    cpp_out_file.write(str_process)

    cpp_out_file.write('''    
    return true;
}
    ''')

    head_out_file.close()
    cpp_out_file.close()



