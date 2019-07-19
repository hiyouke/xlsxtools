#!/usr/bin/python
# coding=utf-8

import types
import sys
import os
from openpyxl import load_workbook


def get_cell_value(a_cell):
    cell_type = type(a_cell.value)
    if cell_type is types.UnicodeType:
        cell_value = a_cell.value.encode('utf-8')
    elif cell_type is types.NoneType:
        cell_value = ''
    else:
        cell_value = str(a_cell.value)
    return cell_value


filename = sys.argv[1].strip()
path_split = filename.split("/")
name_split = path_split[len(path_split) - 1].split(".")
excel_name = name_split[0].capitalize()

configPath = "./Config"
if not os.path.isdir(configPath):
    os.makedirs(configPath)

wb = load_workbook(filename)
sheet_names = wb.get_sheet_names()
listSheet = wb.get_sheet_by_name(name=sheet_names[0])  # 读第一个sheet
col = listSheet.get_highest_column()
row = listSheet.get_highest_row()
for i in range(row):
    cell = listSheet.cell(column=0, row=i)
    value = get_cell_value(cell)
    sheet_name = value.capitalize()
    dataSheet = wb.get_sheet_by_name(value)
    data_col = dataSheet.get_highest_column()
    data_row = dataSheet.get_highest_row()

    # ====================================================================
    # 输出资源文件
    rcs_path = configPath + "/" + excel_name + "_" + sheet_name + ".txt"  # 资源路径
    rcs_file = open(rcs_path, "w")
    for r in range(data_row):
        out = ""
        for c in range(data_col):
            data_cell = dataSheet.cell(column=c, row=r)
            data_value = get_cell_value(data_cell)
            if c != 0:
                out += "#"
            out += data_value
        rcs_file.write(out + "\n")  # os.linesep
    rcs_file.close()
